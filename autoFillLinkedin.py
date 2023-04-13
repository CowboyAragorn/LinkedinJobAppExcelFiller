#! /usr/bin/python3
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from dotenv import load_dotenv
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import date
import traceback


def deleteLastRow():
    try:
        cwd = os.getcwd()
        book = load_workbook(filename=cwd+'/Job_Search_Log.xlsx')
        ws = book.active
        rowToDelete = int(ws.max_row)
        if rowToDelete == 0:
            print('1.Sorry, can\'t delete the first row')
            return
        print(rowToDelete)
        #TODO: format this to display the values being deleted
        print(ws[rowToDelete][0].value)
        ws.delete_rows(rowToDelete)
        print('row ' + str(rowToDelete) + ' deleted.')
        book.save(cwd+'/Job_Search_Log.xlsx')
    except:
        traceback.print_exc()
        print('1.An error occured, nothing has been deleted.')

def headlessBrowse():
#TODO make headless an option
    print('1.To SAVE new job: Right click the name of a job on linkedin\'s job page, copy the link, paste it here, and hit enter:')
    print('')
    url = input()
    # runs selenium browser invisibly so you can just operate from CLI
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--allow-running-insecure-content')
    browser = webdriver.Chrome(options=chrome_options)
    scrapeAndLoad(browser, url)

def debugBrowse():
    print('1.To SAVE new job: Right click the name of a job on linkedin\'s job page, copy the link, paste it here, and hit enter:')
    print('2. You will see the webscrape process in action to debug - the load times will take longer and each failure will show their traceback message')
    print('')
    url = input()
    browser = webdriver.Chrome()
    scrapeAndLoad(browser, url, 'debug')

def scrapeAndLoad(browser, url, type='headless'):
    print('Loading Linkedin...')
    browser.get(url)
    browser.maximize_window()
    if type == 'debug':
        time.sleep(10)
    else:
        time.sleep(3)

    print('Navigating to sign in page...')
    #first sign in click on linkedin
    signInBtn1 = browser.find_element(By.XPATH, '/html/body/div[4]/a[1]')
    signInBtn1.click()
    if type == 'debug':
        time.sleep(10)
    else:
        time.sleep(3)

    print('Inputting credentials...')
    #second sign in click on linkedin
    cwd = os.getcwd()
    load_dotenv(cwd+'/secrets.env')
    username = str(os.getenv('USERNME'))
    password = str(os.getenv('PASSWORD'))
    usernameInput = browser.find_element(By.XPATH, '/html/body/div/main/div[2]/div[1]/form/div[1]/input')
    passwordInput = browser.find_element(By.XPATH, '/html/body/div/main/div[2]/div[1]/form/div[2]/input')
    signInBtn2 = browser.find_element(By.XPATH, '/html/body/div/main/div[2]/div[1]/form/div[3]/button')

    #TODO: Find a way to do this async based on page load, not just waiting arbitrary amount of time

    usernameInput.send_keys(username)
    passwordInput.send_keys(password)
    signInBtn2.click()
    #time.sleep(20)
    if type == 'debug':
        time.sleep(15)
    else:
        time.sleep(5)

    print('Login Successful!')
    #If this was a larger project, make a dict for each of the values keeping info like cell, innerhtml so that you don't have to manually update when making changes
    #create excel if it doesn't exists
    if not (os.path.isfile(cwd+'/Job_Search_Log.xlsx')):
        print('Creating workbook...')
        wb = openpyxl.Workbook()
        wb_name = 'Job_Search_Log.xlsx'
        ws = wb.active
        ws.title ='Job_Search_Log'
        ws['A1'] = 'Name'
        ws['B1'] = 'Company'
        ws['C1'] = 'Date'
        ws['D1'] = 'Location'
        ws['E1'] = 'Link'
        ws['F1'] = 'Salary'
        ws['G1'] = 'Workplace Type'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20
        ws.row_dimensions[1].height = 15
        cell = ws[('A1'):('G1')]
        for i in cell[0]:
            i.fill = PatternFill(start_color='00008000', end_color='00008000', fill_type='solid')
            i.font = Font(bold=True, size=13, color='00FFFFFF')

        ws.freeze_panes = "A2"
        wb.save(cwd+'/Job_Search_Log.xlsx')
        wb.close()

    #Write new info to workbook
    editBook = load_workbook(filename=cwd+'/Job_Search_Log.xlsx')
    ws = editBook.active
    rowToWrite = str(ws.max_row + 1)
    print('Writing to row ' + rowToWrite + '...')

    #color even rows
    if int(rowToWrite)%2 == 0:
        cell = ws[('A'+rowToWrite):('G'+rowToWrite)]
        for i in cell[0]:
            i.fill = PatternFill(start_color='0099CC00', end_color='0099CC00', fill_type='solid')


    #TODO: Perhaps switch link to jobs specifc link

    #NAME
    try:
        name = browser.find_element(By.XPATH, '/html/body/div[5]/div[3]/div/div[1]/div[1]/div/div[1]/div/div/div[1]/h1')
        #print(name.text)
        ws['A' + rowToWrite] = str(name.text)
    except: 
        if type =='debug':
            traceback.print_exc()
        print('name failed')
    #COMPANY
    try:
        company = browser.find_element(By.XPATH,'/html/body/div[5]/div[3]/div/div[1]/div[1]/div/div[1]/div/div/div[1]/div[1]/span[1]/span[1]/a' )
        #print(company.text)
        ws['B' + rowToWrite] = str(company.text)
    except: 
        if type =='debug':
            traceback.print_exc()
        print('company failed')
    #DATE
    try:
        today = date.today()
        dateFormatted = today.strftime("%B %d, %Y")
        ws['C' + rowToWrite] = str(dateFormatted)
    except: 
        if type =='debug':
            traceback.print_exc()
        print('Date failed')
    #LOCATION
    try:
        location = browser.find_element(By.XPATH, '/html/body/div[5]/div[3]/div/div[1]/div[1]/div/div[1]/div/div/div[1]/div[1]/span[1]/span[2]')
        #print(location.text)
        ws['D' + rowToWrite] = str(location.text)
    except: 
        if type =='debug':
            traceback.print_exc()
        print('location failed')
    #LINK
    try:
        # postingLink = browser.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[4]/div/div/main/div/section[2]/div/div[2]/div[1]/div/div[1]/div/div[1]/div[1]/a')
        # postingLink = postingLink.get_attribute('href')
        # ws['E' + rowToWrite] = str(postingLink)
        #just take url from original user input
        ws['E' + rowToWrite] = str(url)
    except: 
        if type =='debug':
            traceback.print_exc()
        print('link failed')
    #SALARY
    try:
        salary = browser.find_element(By.XPATH, '/html/body/div[5]/div[3]/div/div[1]/div[1]/div/div[6]/div[1]/div/div[2]/p')
        #print(postingLink)
        ws['F' + rowToWrite] = str(salary.text)
    except:
        if type =='debug':
            traceback.print_exc()
        print('No Salary Available')
    #WORKPLACE TYPE
    try:
        remote = browser.find_element(By.XPATH, '/html/body/div[5]/div[3]/div/div[1]/div[1]/div/div[1]/div[1]/div/div[1]/div[1]/span[1]/span[3]')
        ws['G' + rowToWrite] = str(remote.text)
    except: 
        if type =='debug':
            traceback.print_exc()
        print('No Remote Info Available')

    editBook.save(cwd+'/Job_Search_Log.xlsx')
    #TODO: tell what you are writing to workbook, ie writing [name,company]
    print('Workbook saved!')


# print('INSTRUCTIONS:')
# print('1.To run normally, hit enter.')
# print('2.To run in DEBUG mode, type debug then hit enter.')
# print('3.To DELETE LAST ENTRY: type del and hit enter.Warning: can not be undone.')
# print('4.To exit program: type exit and hit enter.')
while True: 
    print('INSTRUCTIONS:')
    print('1.To run normally, hit enter.')
    print('2.To run in DEBUG mode, type debug then hit enter.')
    print('3.To DELETE LAST ENTRY: type del and hit enter.Warning: can not be undone.')
    print('4.To exit program: type exit and hit enter.')
    print('')
    mode = input()
    if mode == '':
        headlessBrowse()
    elif mode == 'debug':
        debugBrowse()
    elif mode == 'del':
        print('1.Are you sure you want to delete the last row? Type y to confirm, type anything else to return to previous menu')
        print('')
        confirm = input()
        if confirm == 'y':
            deleteLastRow()
        else: 
            continue
    elif mode == 'exit':
        exit()
    else:
        print('1.Invalid input, try one of the options again. All entries must be lowercase.')
        print('')
        continue

    print('1.Action completed. would you like to run the program again? y/n')
    print('')
    mode2 = input()
    if mode2 == 'y':
        continue
    else:
        print('1.Goodbye!')
        exit()

    
